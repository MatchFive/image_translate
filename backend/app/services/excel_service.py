"""
Excel 处理服务

负责读取/写入 Excel 文件，提取图片和文本。
"""

import io
import uuid
import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from PIL import Image

from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()


class ExcelItem:
    """Excel 中的一行数据"""

    def __init__(self, row_index: int, image_path: str, target_text: str):
        self.row_index = row_index
        self.image_path = image_path
        self.target_text = target_text


class ExcelProcessor:
    """Excel 文件处理器"""

    def __init__(self, task_id: str):
        self.task_id = task_id
        self.task_dir = settings.UPLOAD_DIR / task_id
        self.task_dir.mkdir(parents=True, exist_ok=True)

    def parse_columns(self, filepath: str | Path) -> list[dict]:
        """
        解析 Excel 列头信息，返回所有列的序号和名称。

        Returns:
            [{"index": 1, "name": "图片"}, {"index": 2, "name": "文本"}, ...]
        """
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        columns = []

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            col_name = str(cell.value).strip() if cell.value else f"第{col_idx}列"
            columns.append({
                "index": col_idx,
                "name": col_name,
            })

        wb.close()
        return columns

    def read_excel(
        self,
        filepath: str | Path,
        image_col: int = 1,
        text_col: int = 2,
    ) -> list[ExcelItem]:
        """
        读取 Excel 文件，提取图片和目标文本。

        Args:
            filepath: Excel 文件路径
            image_col: 图片列序号（1-based），默认第1列
            text_col: 文字列序号（1-based），默认第2列

        Returns:
            ExcelItem 列表

        处理逻辑：
        1. 从指定 image_col 列提取嵌入的图片
        2. 从指定 text_col 列读取目标替换文本
        3. 将图片保存到任务目录
        """
        wb = load_workbook(filepath)
        ws = wb.active
        items = []

        # 提取所有嵌入图片，建立 (row, col) -> 图片文件路径 的映射
        image_map = self._extract_images(ws)

        for row_idx in range(2, ws.max_row + 1):
            # 读取指定文字列的目标文本
            target_text_cell = ws.cell(row=row_idx, column=text_col)
            target_text = str(target_text_cell.value).strip() if target_text_cell.value else ""

            if not target_text:
                logger.warning(f"Row {row_idx}: 跳过，目标文本为空")
                continue

            # 查找该行指定图片列对应的图片
            image_path = image_map.get((row_idx, image_col))
            if not image_path:
                # 降级：只按行号查找（兼容旧行为）
                image_path = image_map.get(row_idx)
            if not image_path:
                logger.warning(f"Row {row_idx}: 跳过，未找到图片")
                continue

            items.append(ExcelItem(
                row_index=row_idx - 2,  # 0-based 索引
                image_path=image_path,
                target_text=target_text,
            ))

        wb.close()
        logger.info(f"Task {self.task_id}: 从 Excel 提取 {len(items)} 个处理项 (图片列={image_col}, 文字列={text_col})")
        return items

    def _extract_images(self, ws) -> dict[int | tuple, str]:
        """
        从工作表中提取所有图片。

        Returns:
            {row_number: image_file_path} 映射（兼容旧逻辑）
            同时也返回 {(row_number, col_number): image_file_path} 映射（支持多列）
        """
        image_map = {}

        for idx, img in enumerate(ws._images):
            try:
                anchor = img.anchor
                row = None
                col = None

                if hasattr(anchor, '_from'):
                    row = anchor._from.row + 1  # 1-based
                    col = getattr(anchor._from, 'col', None)
                    if col is not None:
                        col += 1  # 1-based
                elif hasattr(anchor, 'row'):
                    row = anchor.row + 1

                if row is None:
                    row = idx + 2
                    logger.warning(f"Image {idx}: 无法定位行号，默认分配到 row {row}")
                if col is None:
                    col = 1
                    logger.warning(f"Image {idx}: 无法定位列号，默认分配到 col {col}")

            except Exception:
                row = idx + 2
                col = 1
                logger.warning(f"Image {idx}: 定位异常，默认分配到 row {row}, col {col}")

            # 保存图片到任务目录
            img_filename = f"original_{row}_{col}.png"
            img_path = self.task_dir / img_filename

            image_data = img._data()
            pil_img = Image.open(io.BytesIO(image_data))
            pil_img.save(str(img_path), "PNG")

            # 同时存储行号映射和(row, col)映射
            image_map[row] = str(img_path)
            image_map[(row, col)] = str(img_path)
            logger.debug(f"Image extracted: row {row}, col {col} -> {img_path}")

        return image_map

    def write_results(
        self,
        source_filepath: str | Path,
        results: dict[int, str],
        output_filepath: str | Path | None = None,
        image_col: int = 1,
    ) -> str:
        """
        将处理结果写回 Excel 结果列（图片列旁边新增一列）。

        Args:
            source_filepath: 原始 Excel 路径
            results: {row_index(0-based): result_image_path} 映射
            output_filepath: 输出文件路径（默认覆盖原文件）
            image_col: 原始图片列序号（1-based），结果将写入该列右侧

        Returns:
            结果文件路径
        """
        source_filepath = Path(source_filepath)
        output_filepath = Path(output_filepath) if output_filepath else source_filepath

        wb = load_workbook(source_filepath)
        ws = wb.active

        result_col = image_col + 2  # 跳过图片列和文字列，结果放在后面
        # 确保结果列有表头
        header_cell = ws.cell(row=1, column=result_col)
        if not header_cell.value:
            header_cell.value = "处理结果"

        for row_idx_0, result_path in results.items():
            excel_row = row_idx_0 + 2  # 转回 Excel 行号（跳过表头）
            if not Path(result_path).exists():
                logger.warning(f"Row {excel_row}: 结果图片不存在 {result_path}")
                continue

            col_letter = chr(ord('A') + result_col - 1)  # 转列字母
            img = XlImage(result_path)
            img.width = min(img.width, 200)
            img.height = min(img.height, 200)
            ws.add_image(img, f"{col_letter}{excel_row}")

        wb.save(str(output_filepath))
        wb.close()
        logger.info(f"Task {self.task_id}: 结果已写入 {output_filepath}")
        return str(output_filepath)
